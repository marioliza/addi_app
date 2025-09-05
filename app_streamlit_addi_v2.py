import io
from io import BytesIO
import zipfile
from pathlib import Path
from typing import Dict, Any, List, Tuple
import os
import unicodedata
import random
import re

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# =========================
# TITLE & AUTH
# =========================
APP_TITLE = "🧩 Herramienta para crear órdenes – Seller Addi"
APP_SUBTITLE = "Carga tu Excel origen y template, mapea columnas y descarga los archivos listos (100 por archivo)."
REQUIRED_PASSWORD = "addi2025*"  # Puedes sobreescribir con st.secrets['APP_PASSWORD']

st.set_page_config(page_title="Seller Addi – Crear Órdenes", layout="wide")
st.title(APP_TITLE)
st.caption(APP_SUBTITLE)

# Password gate (simple)
try:
    pw = st.secrets["APP_PASSWORD"]  # usa secrets si existe
except Exception:
    pw = os.environ.get("APP_PASSWORD", REQUIRED_PASSWORD)  # o variable de entorno o fallback
if "auth_ok" not in st.session_state:
    st.session_state.auth_ok = False

if not st.session_state.auth_ok:
    st.subheader("🔒 Acceso")
    typed = st.text_input("Contraseña", type="password", placeholder="Ingresa la contraseña")
    if st.button("Entrar", type="primary"):
        if typed == pw:
            st.session_state.auth_ok = True
            st.success("Acceso concedido.")
            st.rerun()
        else:
            st.error("Contraseña incorrecta.")
    st.stop()

# =========================
# PROGRESS
# =========================
class ProgressTracker:
    def __init__(self, total_rows, label="Procesando"):
        self.total = max(int(total_rows), 1)
        self.done = 0
        self.pbar = st.progress(0, text=f"{label} 0% (0/{self.total})")
    def add(self, n=1, label="Procesando"):
        self.done += int(n)
        if self.done > self.total:
            self.done = self.total
        frac = self.done / self.total
        pct = int(frac * 100)
        self.pbar.progress(frac, text=f"{label} {pct}% ({self.done}/{self.total})")
    def finish(self, label="Completado"):
        self.pbar.progress(1.0, text=f"{label} 100% ({self.total}/{self.total})")

# =========================
# CONFIG (ajustable en código)
# =========================
WAREHOUSES = [
    {"label": "Bogotá #2 - Montevideo", "city": "Bogotá"},
    {"label": "Medellin #2 - Sabaneta Mayorca", "city": "Medellin"},  # SIN tilde
]

# ===== Normalización / utilidades =====
def _norm(s: str) -> str:
    s = str(s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return s

def _norm_hard(s: str) -> str:
    return re.sub(r"\s+", " ", _norm(s)).strip()

def _slugify_no_spaces(s: str) -> str:
    # minúscula, sin acentos, quitar cualquier cosa que no sea a-z0-9, quitar espacios
    s0 = _norm(s)
    s1 = re.sub(r"[^a-z0-9]+", "", s0)  # solo alfanumérico
    return s1

def make_external_order_slug(brand: str, empresa: str) -> str:
    b = _slugify_no_spaces(brand or "")
    e = _slugify_no_spaces(empresa or "")
    return f"{b}-{e}".strip("-")

def _get_wh_label_for_city(hub_city_norm: str) -> str:
    for wh in WAREHOUSES:
        if _norm(wh.get("city", "")) == hub_city_norm:
            return wh["label"]
    return WAREHOUSES[0]["label"]

# =========================
# NUEVO CITY_TO_HUB (optimizado por menor tiempo terrestre)
# =========================
CITY_TO_HUB = {
    # Área Metropolitana de Medellín + cercanías (ya estaban OK)
    "medellin": "medellin", "medellín": "medellin", "itagui": "medellin", "itagüi": "medellin",
    "envigado": "medellin", "sabaneta": "medellin", "bello": "medellin", "la estrella": "medellin",
    "caldas": "medellin", "girardota": "medellin", "copacabana": "medellin",
    # Oriente cercano + Urabá (OK)
    "rionegro": "medellin", "marinilla": "medellin", "la ceja": "medellin", "guarne": "medellin",
    "carmen de viboral": "medellin", "el retiro": "medellin",
    "turbo": "medellin", "apartado": "medellin", "apartadó": "medellin", "necocli": "medellin", "necoclí": "medellin",

    # Eje cafetero (mejor Medellín)
    "pereira": "medellin", "dosquebradas": "medellin", "santa rosa de cabal": "medellin",
    "manizales": "medellin", "villamaria": "medellin", "villamaría": "medellin",
    "armenia": "medellin", "circasia": "medellin", "montenegro": "medellin", "quimbaya": "medellin",
    "la tebaida": "medellin", "filandia": "medellin",

    # Norte del Valle (mejor Medellín)
    "cartago": "medellin", "roldanillo": "medellin", "zarzal": "medellin", "sevilla": "medellin",
    "la union": "medellin", "la unión": "medellin",

    # Costa Caribe (cambiar a Medellín)
    "barranquilla": "medellin", "cartagena": "medellin", "santa marta": "medellin", "riohacha": "medellin",
    "valledupar": "medellin", "monteria": "medellin", "montería": "medellin",
    "sincelejo": "medellin", "magangue": "medellin", "magangué": "medellin",
    "corozal": "medellin", "tolu": "medellin", "tolú": "medellin",
    "galapa": "medellin", "malambo": "medellin", "baranoa": "medellin", "puerto colombia": "medellin",
    "san onofre": "medellin", "turbaco": "medellin", "mahates": "medellin",
    "el banco": "medellin", "aracataca": "medellin", "fundacion": "medellin", "fundación": "medellin",
    "cienaga": "medellin", "ciénaga": "medellin", "dibulla": "medellin", "uribia": "medellin", "maicao": "medellin",
    "santa rosa del sur": "medellin", "el carmen de bolivar": "medellin", "el carmen de bolívar": "medellin",

    # Santander – split
    # Área metropolitana de Bucaramanga → Medellín
    "bucaramanga": "medellin", "floridablanca": "medellin", "piedecuesta": "medellin",
    "giron": "medellin", "girón": "medellin", "lebrija": "medellin",
    # Provincia Guanentá etc. → Bogotá
    "san gil": "bogota", "curiti": "bogota", "curití": "bogota",
    "el socorro": "bogota", "barbosa": "bogota",

    # Norte de Santander – split
    # Capital y área cercana → Medellín (ligeramente menor tiempo)
    "cucuta": "medellin", "cúcuta": "medellin", "el zulia": "medellin",
    # Corredor Pamplona/Chinácota/Toledo → Bogotá
    "pamplona": "bogota", "chinacota": "bogota", "chinácota": "bogota", "toledo": "bogota",
    "abrego": "bogota", "ábrego": "bogota", "sardinata": "bogota",

    # Cundinamarca/Sabana (Bogotá)
    "bogota": "bogota", "bogotá": "bogota", "soacha": "bogota", "funza": "bogota", "mosquera": "bogota",
    "madrid": "bogota", "chia": "bogota", "chía": "bogota", "zipaquira": "bogota", "zipaquirá": "bogota",
    "cajica": "bogota", "cajicá": "bogota", "tocancipa": "bogota", "tocancipá": "bogota",
    "cota": "bogota", "la calera": "bogota",

    # Boyacá (Bogotá)
    "tunja": "bogota", "paipa": "bogota", "villa de leyva": "bogota",
    "chiquinquira": "bogota", "chiquinquirá": "bogota", "samaca": "bogota", "samacá": "bogota",

    # Tolima (Bogotá)
    "ibague": "bogota", "ibagué": "bogota", "espinal": "bogota", "melgar": "bogota",
    "honda": "bogota", "rovira": "bogota", "lerida": "bogota", "lérida": "bogota",
    "mariquita": "bogota", "chaparral": "bogota", "icononzo": "bogota", "fresno": "bogota",
    "tocaima": "bogota", "purificacion": "bogota", "purificación": "bogota",
    "saldaña": "bogota", "villahermosa": "bogota",

    # Huila (Bogotá)
    "neiva": "bogota", "pitalito": "bogota", "garzon": "bogota", "garzón": "bogota",
    "hobo": "bogota", "campoalegre": "bogota", "tarqui": "bogota", "palestina": "bogota", "la plata": "bogota",

    # Meta / Llanos (Bogotá)
    "villavicencio": "bogota", "acacias": "bogota", "acacías": "bogota",
    "granada": "bogota", "cumaral": "bogota", "san martin": "bogota", "san martín": "bogota",
    "restrepo": "bogota", "vista hermosa": "bogota", "puerto lopez": "bogota", "puerto lópez": "bogota",

    # Casanare / Arauca (Bogotá)
    "yopal": "bogota", "tauramena": "bogota", "aguazul": "bogota", "paz de ariporo": "bogota",
    "arauca": "bogota", "saravena": "bogota", "arauquita": "bogota",

    # Caquetá / Putumayo / Guaviare / Amazonas (Bogotá)
    "florencia": "bogota", "san vicente del caguan": "bogota", "san vicente del caguán": "bogota",
    "cartagena del chaira": "bogota", "cartagena del chairá": "bogota",
    "el doncello": "bogota", "el pital": "bogota",
    "mocoa": "bogota", "orito": "bogota", "puerto asis": "bogota", "puerto asís": "bogota", "sibundoy": "bogota",
    "san jose del guaviare": "bogota", "san josé del guaviare": "bogota",
    "el retorno": "bogota",
    "leticia": "bogota", "puerto nariño": "bogota",

    # Nariño (ligeramente mejor Medellín)
    "pasto": "medellin", "ipiales": "medellin", "tuquerres": "medellin", "túquerres": "medellin",
    "cumbal": "medellin", "tumaco": "medellin",

    # Valle del Cauca (recomiendo Medellín)
    "cali": "medellin", "yumbo": "medellin", "buga": "medellin",
    "palmira": "medellin", "el cerrito": "medellin", "florida": "medellin", "pradera": "medellin",
}

# =========================
# NUEVO DEPT_TO_HUB
# (usa departamentos como fallback; las ciudades arriba prevalecen)
# =========================
DEPT_TO_HUB = {
    # Medellín por Caribe y Eje Cafetero/Norte del Valle/Valle/Nariño
    "antioquia": "medellin",
    "cordoba": "medellin", "córdoba": "medellin",
    "atlantico": "medellin", "atlántico": "medellin",
    "bolivar": "medellin", "bolívar": "medellin",
    "magdalena": "medellin", "cesar": "medellin", "sucre": "medellin", "la guajira": "medellin",
    "risaralda": "medellin", "quindio": "medellin", "quindío": "medellin", "caldas": "medellin",
    "valle del cauca": "medellin",
    "narino": "medellin", "nariño": "medellin",

    # Bogotá como fallback para el resto del centro-oriente y suroriente
    "cundinamarca": "bogota", "bogota, d.c.": "bogota", "bogota d.c.": "bogota", "bogotá d.c.": "bogota",
    "boyaca": "bogota", "boyacá": "bogota",
    "tolima": "bogota", "huila": "bogota", "meta": "bogota",
    "santander": "bogota",               # (con ciudades arriba que override a Medellín)
    "norte de santander": "bogota",      # (con split por ciudades arriba)
    "arauca": "bogota", "casanare": "bogota",
    "caqueta": "bogota", "caquetá": "bogota", "putumayo": "bogota",
    "guaviare": "bogota", "amazonas": "bogota",
}


KEYWORDS_MEDELLIN = ["medellin", "sabaneta", "itagui", "envigado", "bello", "antioquia", "uraba", "turbo", "apartado", "necocli"]
KEYWORDS_BOGOTA   = ["bogota", "cundinamarca", "sabana", "zipaquira", "chia", "tocancipa", "boyaca", "santander", "tolima", "meta", "huila", "llanos"]

def assign_bodega_by_city(row: pd.Series) -> str:
    city_val = _norm(row.get("Ciudad", ""))
    dept_val = _norm(row.get("Departamento", ""))
    hub = CITY_TO_HUB.get(city_val)
    if hub: return _get_wh_label_for_city(hub)
    hub = DEPT_TO_HUB.get(dept_val)
    if hub: return _get_wh_label_for_city(hub)
    if any(k in city_val or k in dept_val for k in KEYWORDS_MEDELLIN): return _get_wh_label_for_city("medellin")
    if any(k in city_val or k in dept_val for k in KEYWORDS_BOGOTA):   return _get_wh_label_for_city("bogota")
    return _get_wh_label_for_city("bogota")

# =========================
# SIDEBAR
# =========================
with st.sidebar:
    st.header("⚙️ Configuración")
    chunk_size = st.number_input("Tamaño por archivo (máx. registros)", min_value=1, value=100, step=1)
    header_row = st.number_input("Fila de encabezados del template", min_value=1, value=1, step=1)
    start_row = st.number_input("Fila inicial de escritura", min_value=1, value=3, step=1)  # Escribir desde A3
    default_prefix = st.text_input("Prefijo del nombre de salida", value="template_part")
    st.caption("La asignación de **Bodega** se hace por ciudad (mapeo) con fallback por departamento. **La Ciudad se mantiene tal cual del origen**.")

col_u1, col_u2 = st.columns(2)

# =========================
# UPLOAD SOURCE
# =========================
with col_u1:
    src_file = st.file_uploader("📥 Excel **origen** (.xlsx)", type=["xlsx"], key="src")
    src_sheet = None
    if src_file:
        try:
            xls = pd.ExcelFile(src_file)
            src_sheet = st.selectbox("Hoja de origen", xls.sheet_names, index=0, key="src_sheet")
            src_df = xls.parse(src_sheet, dtype=object)
            src_df.columns = [str(c).strip() for c in src_df.columns]
            st.success(f"Origen cargado. Filas: {len(src_df):,}. Columnas: {len(src_df.columns)}")

            # =========================
            # LIMPIEZAS / FORMATEO
            # =========================
            phones_autofilled = 0
            emails_cleared = 0

            # 1) Teléfonos vacíos → generar número aleatorio válido (10 dígitos, inicia en 3)
            if "Celular" in src_df.columns:
                def _random_phone():
                    return "3" + "".join(str(random.randint(0, 9)) for _ in range(9))
                src_df["Celular"] = src_df["Celular"].fillna("").astype(str)
                empties = src_df["Celular"].str.strip() == ""
                phones_autofilled = int(empties.sum())
                if phones_autofilled > 0:
                    src_df.loc[empties, "Celular"] = [_random_phone() for _ in range(phones_autofilled)]

            # 2) Correos: solo gmail/hotmail en minúscula, otros → BLANCO
            if "Correo electrónico" in src_df.columns:
                src_df["Correo electrónico"] = src_df["Correo electrónico"].fillna("").astype(str).str.lower()
                mask_valid = (
                    src_df["Correo electrónico"].str.endswith("@gmail.com")
                    | src_df["Correo electrónico"].str.endswith("@hotmail.com")
                )
                emails_cleared = int((~mask_valid & src_df["Correo electrónico"].ne("")).sum())
                src_df.loc[~mask_valid, "Correo electrónico"] = ""

            # 3) Generar "Número de orden externo" = brand-slug + "-" + empresa (sin espacios/acentos)
            if "Brand Slug" in src_df.columns and "Nombre de la empresa" in src_df.columns:
                src_df["Número de orden externo"] = src_df.apply(
                    lambda r: make_external_order_slug(r.get("Brand Slug", ""), r.get("Nombre de la empresa", "")),
                    axis=1
                )

            # Guardar métricas parciales
            st.session_state._metrics = {
                "phones_autofilled": phones_autofilled,
                "emails_cleared": emails_cleared,
            }

            with st.expander("Vista previa origen (ya formateado)", expanded=False):
                st.dataframe(src_df.head(20))
        except Exception as e:
            st.error(f"Error leyendo origen: {e}")
            src_df = None
    else:
        src_df = None

# =========================
# UPLOAD TEMPLATE
# =========================
with col_u2:
    tmpl_file = st.file_uploader("📥 Excel **template** (.xlsx)", type=["xlsx"], key="tmpl")
    target_sheet = None
    headers: List[str] = []
    header_index: Dict[str, int] = {}
    header_positions: Dict[str, List[int]] = {}
    if tmpl_file:
        try:
            tmpl_bytes = tmpl_file.getvalue()
            wb = load_workbook(filename=BytesIO(tmpl_bytes), data_only=True)
            target_sheet = st.selectbox("Hoja del template", wb.sheetnames, index=0, key="tmpl_sheet")
            ws = wb[target_sheet]
            headers = []
            header_positions = {}
            for idx, cell in enumerate(ws[header_row], start=1):
                v = cell.value
                if v is None:
                    headers.append("")
                else:
                    name = str(v).strip()
                    headers.append(name)
                    header_positions.setdefault(name, []).append(idx)
            header_index = {name: i+1 for i, name in enumerate(headers) if name}
            st.success(f"Template cargado. Hoja '{target_sheet}'. Encabezados encontrados: {len(header_index)}")
            with st.expander(f"Encabezados del template (fila {header_row})", expanded=False):
                st.write([h for h in headers])
        except Exception as e:
            st.error(f"Error leyendo template: {e}")
            tmpl_bytes = None
    else:
        tmpl_bytes = None

st.markdown("---")
st.subheader("🧭 Mapeo de columnas (destino → origen / constante)")

# Mapeo por defecto (ajustado: 'Número de orden externo' ahora viene del origen ya calculado)
preset_mapping = {
    "Plantilla": {"mode": "template_name"},
    "Número de orden externo": {"mode": "source", "source_col": "Número de orden externo"},  # ← slug brand-empresa
    "Nombre completo del comprador": {"mode": "source", "source_col": "Nombre completo"},
    # "Indicativo": se fuerza abajo en col C con 57; otras en blanco
    "Teléfono de contacto": {"mode": "source", "source_col": "Celular"},
    "Correo electrónico": {"mode": "source", "source_col": "Correo electrónico"},
    "Tipo de empacado": {"mode": "const", "const_value": "Estandar"},
    "Igual al comprador": {"mode": "const", "const_value": "SI"},
    "Dirección": {"mode": "source", "source_col": "Dirección"},
    "Ciudad": {"mode": "source", "source_col": "Ciudad"},  # NO se cambia
    "Región": {"mode": "source", "source_col": "Departamento"},
    "País": {"mode": "const", "const_value": "Colombia"},
    "Método de envío": {"mode": "const", "const_value": "Estándar (Local y Nacional)"},
    "Tipo de recaudo": {"mode": "const", "const_value": "NO APLICA"},
    "SKU o Código Melonn del producto": {"mode": "source", "source_col": "Referencia"},
    "Cantidad": {"mode": "source", "source_col": "Número de tiendas"},
    # Bodega/CEDIS se llenará automáticamente por reglas
}

if "mapping_state" not in st.session_state:
    st.session_state.mapping_state = {}

def draw_mapping_ui(headers: List[str], src_cols: List[str]) -> Dict[str, Any]:
    mapping: Dict[str, Any] = {}
    modes = ["(no escribir)", "source", "const", "template_name", "source_filename"]
    col_left, col_right = st.columns(2)
    with col_left:
        st.caption("Destino (Template)")
    with col_right:
        st.caption("Origen / Valor")

    for dest in headers:
        if not dest:
            continue
        prev = st.session_state.mapping_state.get(dest, {})
        default_mode = "(no escribir)"
        default_source = src_cols[0] if src_cols else ""
        default_const = ""
        if dest in preset_mapping:
            pm = preset_mapping[dest]
            default_mode = pm.get("mode", default_mode)
            default_source = pm.get("source_col", default_source)
            default_const = str(pm.get("const_value", default_const))
        if prev:
            default_mode = prev.get("mode", default_mode)
            default_source = prev.get("source_col", default_source)
            default_const = str(prev.get("const_value", default_const))

        lock_dest = dest in ("Bodega", "CEDIS de origen")  # se calculan automático
        c1, c2 = st.columns([1, 2])
        with c1:
            mode = st.selectbox(
                f"{dest}",
                options=modes,
                index=modes.index(default_mode) if default_mode in modes else 0,
                key=f"mode::{dest}",
                disabled=lock_dest
            )
        with c2:
            if mode == "source":
                source_col = st.selectbox(
                    f"Columna origen para '{dest}'",
                    options=src_cols,
                    index=src_cols.index(default_source) if default_source in src_cols else 0 if src_cols else 0,
                    key=f"src::{dest}",
                    disabled=lock_dest
                )
                mapping[dest] = {"mode": "source", "source_col": source_col}
            elif mode == "const":
                const_val = st.text_input(
                    f"Valor fijo para '{dest}'",
                    value=default_const,
                    key=f"const::{dest}",
                    disabled=lock_dest
                )
                mapping[dest] = {"mode": "const", "const_value": const_val}
            elif mode == "template_name":
                st.text("(escribirá el nombre del template)")
                mapping[dest] = {"mode": "template_name"}
            elif mode == "source_filename":
                st.text("(escribirá el nombre del archivo origen)")
                mapping[dest] = {"mode": "source_filename"}
            else:
                st.text("—")
                mapping[dest] = {"mode": "(no escribir)"}

        st.session_state.mapping_state[dest] = mapping[dest]

    return mapping

src_cols_list = list(src_df.columns) if src_df is not None else []
mapping = draw_mapping_ui(list(header_index.keys()) if header_index else [], src_cols_list)

# =========================
# HELPERS
# =========================
def resolve_value(spec: Dict[str, Any], row: pd.Series, template_name: str, source_name: str):
    mode = spec.get("mode", "(no escribir)")
    if mode == "source":
        col = spec.get("source_col", "")
        return row.get(col, None)
    elif mode == "const":
        val = spec.get("const_value", "")
        try:
            f = float(val)
            if f.is_integer():
                return int(f)
            return f
        except Exception:
            return val
    elif mode == "template_name":
        return template_name
    elif mode == "source_filename":
        return source_name
    else:
        return None

def fill_one_chunk(
    tmpl_bytes: bytes,
    target_sheet: str,
    header_index: Dict[str, int],
    header_positions: Dict[str, List[int]],
    start_row: int,
    chunk_df: pd.DataFrame,
    mapping: Dict[str, Any],
    template_name: str,
    source_name: str,
    prog: ProgressTracker,
) -> Tuple[bytes, Dict[str, int]]:
    wb = load_workbook(filename=BytesIO(tmpl_bytes))
    if target_sheet not in wb.sheetnames:
        raise KeyError(f"La hoja '{target_sheet}' no existe en el template.")
    ws = wb[target_sheet]

    stats = {"rows": 0, "nw_written": 0, "no_dest_bodega": 0}

    # Detectar columna destino para bodega (prioriza 'Bodega', luego 'CEDIS de origen')
    dest_bodega = None
    if "Bodega" in header_index:
        dest_bodega = "Bodega"
    elif "CEDIS de origen" in header_index:
        dest_bodega = "CEDIS de origen"

    for r_offset, (_, row) in enumerate(chunk_df.iterrows()):
        row_idx = start_row + r_offset

        # 1) Mapeo normal
        for dest, spec in mapping.items():
            if spec.get("mode") == "(no escribir)":
                continue
            if dest not in header_index:
                continue
            c_idx = header_index[dest]
            value = resolve_value(spec, row, template_name, source_name)
            ws.cell(row=row_idx, column=c_idx, value=value)

        # 2) Bodega automática
        b_label = assign_bodega_by_city(row)
        if dest_bodega and dest_bodega in header_index:
            c_bod = header_index[dest_bodega]
            ws.cell(row=row_idx, column=c_bod, value=b_label)
            stats["nw_written"] += 1
        else:
            stats["no_dest_bodega"] += 1

        # 3) Indicativo: solo columna C con 57; otras 'Indicativo' vacías
        indic_idxs = []
        for name, idxs in header_positions.items():
            if str(name).strip().lower() == "indicativo":
                indic_idxs.extend(idxs)
        if indic_idxs:
            keep_idx = 3 if 3 in indic_idxs else indic_idxs[0]
            for c_idx in indic_idxs:
                if c_idx == keep_idx:
                    ws.cell(row=row_idx, column=c_idx, value=57)
                else:
                    ws.cell(row=row_idx, column=c_idx, value=None)

        stats["rows"] += 1
        try:
            prog.add(1, label="Procesando")
        except Exception:
            pass

    out_buf = BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    return out_buf.getvalue(), stats

# =========================
# CONSOLIDACIÓN: 1 registro por (Brand Slug, Nombre de la empresa)
# Suma Número de tiendas con CAP=4 y fija "Número de orden externo" = brand-empresa
# =========================
def consolidate_by_brand_company(df: pd.DataFrame) -> pd.DataFrame:
    BRAND = "Brand Slug"
    EMP   = "Nombre de la empresa"
    QTY   = "Número de tiendas"

    if QTY not in df.columns:
        st.warning("No se encontró la columna 'Número de tiendas' en el origen. No se consolidará.")
        return df
    if BRAND not in df.columns or EMP not in df.columns:
        st.warning("Faltan columnas para la llave (Brand Slug, Nombre de la empresa). No se consolidará.")
        return df

    # Generar/asegurar "Número de orden externo" en TODO el DF (siempre que se pueda)
    df_in = df.copy()
    df_in["Número de orden externo"] = df_in.apply(
        lambda r: make_external_order_slug(r.get(BRAND, ""), r.get(EMP, "")),
        axis=1
    )

    # Llaves normalizadas para agrupar
    df_in["__b__"] = df_in[BRAND].astype(str).map(_norm_hard)
    df_in["__e__"] = df_in[EMP].astype(str).map(_norm_hard)

    rows_out = []
    groups = 0
    removed = 0
    total_qty = 0

    for _, g in df_in.groupby(["__b__", "__e__"], dropna=False):
        groups += 1
        qty_sum = pd.to_numeric(g[QTY], errors="coerce").fillna(0).astype(int).sum()
        qty_cap = min(int(qty_sum), 4)  # CAP=4 por llave brand+empresa
        total_qty += qty_cap

        rep = g.iloc[0].copy()
        rep[QTY] = qty_cap

        # "Número de orden externo" fijo como brand-empresa (recalcular por si acaso)
        rep["Número de orden externo"] = make_external_order_slug(rep.get(BRAND, ""), rep.get(EMP, ""))

        # limpiar auxiliares
        aux_cols = [c for c in rep.index if str(c).startswith("__")]
        if aux_cols:
            rep = rep.drop(labels=aux_cols)

        rows_out.append(rep)
        removed += (len(g) - 1)

    out = pd.DataFrame(rows_out)

    # Métricas
    st.session_state._metrics = {
        **st.session_state.get("_metrics", {}),
        "brand_company_groups": groups,
        "brand_company_removed": removed,
        "cap_per_group": 4,
        "total_qty_after_cap": int(total_qty),
    }

    st.info(f"Consolidación (Brand Slug + Empresa): grupos={groups:,}, filas eliminadas={removed:,}, tope=4 por llave.")
    return out

# =========================
# GENERATE
# =========================
st.markdown("---")
st.subheader("🚀 Generar archivos")

do_run = st.button(
    "Procesar y generar ZIP",
    type="primary",
    disabled=(src_df is None or tmpl_bytes is None or not header_index)
)

if do_run:
    try:
        if src_df is None or tmpl_bytes is None:
            st.stop()
        if len(src_df) == 0:
            st.warning("El origen no tiene filas para procesar.")
            st.stop()

        # >>>> CONSOLIDACIÓN JUSTO ANTES DE ESCRIBIR A EXCEL <<<<
        # 1 registro por (Brand Slug, Nombre de la empresa), sumando y CAP=4; además fija "Número de orden externo".
        src_df = consolidate_by_brand_company(src_df)

        total = len(src_df)
        num_parts = (total + chunk_size - 1) // chunk_size
        st.info(f"Total filas (tras consolidación): {total}. Tamaño de bloque: {chunk_size}. Partes a generar: {num_parts}.")

        template_stem = Path(getattr(tmpl_file, "name", "template.xlsx")).stem
        source_stem = Path(getattr(src_file, "name", "origen.xlsx")).stem

        zip_buf = BytesIO()
        agg = {"rows": 0, "nw_written": 0, "no_dest_bodega": 0}
        prog = ProgressTracker(total_rows=total, label="Procesando")

        with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for i in range(num_parts):
                start = i * chunk_size
                end = min(start + chunk_size, total)
                chunk = src_df.iloc[start:end].copy()

                out_xlsx, stats = fill_one_chunk(
                    tmpl_bytes=tmpl_bytes,
                    target_sheet=target_sheet,
                    header_index=header_index,
                    header_positions=header_positions,
                    start_row=int(start_row),
                    chunk_df=chunk,
                    mapping=mapping,
                    template_name=template_stem,
                    source_name=source_stem,
                    prog=prog,
                )

                for k in agg:
                    agg[k] += stats.get(k, 0)

                part_name = f"{default_prefix}{i+1:02d}.xlsx"
                zf.writestr(part_name, out_xlsx)

        try:
            prog.finish(label="Completado")
        except Exception:
            pass

        zip_buf.seek(0)
        st.success("¡Listo! ZIP generado con registros consolidados por Brand+Empresa y “Número de orden externo” fijo.")
        st.download_button(
            "⬇️ Descargar ZIP",
            data=zip_buf.getvalue(),
            file_name=f"{default_prefix}_lotes.zip",
            mime="application/zip",
        )

        # === Resumen final (incluye métricas de limpieza y consolidación) ===
        metrics = st.session_state.get("_metrics", {})
        with st.expander("Resumen de procesamiento", expanded=True):
            st.write({
                **agg,
                **metrics
            })
    except Exception as e:
        st.error(f"ERROR: {e}")

with st.expander("📝 Notas", expanded=False):
    st.markdown("""
    - **Bodega**: se escribe en 'Bodega' (si existe) o 'CEDIS de origen'. Se decide por mapeo de **ciudad** con fallback por **departamento**/keywords.
    - **Ciudad**: se mantiene exactamente como viene del **origen**.
    - **Indicativo**: solo se llena la **columna C** (si el encabezado es 'Indicativo') con **57**; otras 'Indicativo' se dejan vacías.
    - **Correos**: solo `@gmail.com` o `@hotmail.com` (minúscula). Otros → **en blanco**.
    - **Teléfonos vacíos**: se autocompletan con un número colombiano válido (10 dígitos iniciando en 3).
    - **Consolidación final**: se agrupa por **(Brand Slug, Nombre de la empresa)**, se **suman** unidades con **tope 4** por llave y se fija **“Número de orden externo”** como `brand-empresa` (minúscula, sin acentos ni espacios).
    - **Escritura**: inicia en **A3** (configurable) y divide en archivos del tamaño elegido.
    """)
